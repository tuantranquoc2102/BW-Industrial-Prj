import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from calendar import month_abbr
from openpyxl.styles import PatternFill, Alignment, Font


def read_months_to_show(file_path):
    df = pd.read_excel(file_path)
    return df['months_to_show'].dropna().astype(int).tolist()

def write_summary_full_trends_filtered(
    file_path, trend1, trend2, months, month_filter_path,
    rule_name="Rule", label1="BS trend", label2="PL trend"
):
    # Đọc các tháng cần hiển thị từ file
    months_to_show = read_months_to_show(month_filter_path)

    # Lọc các tháng và xu hướng tương ứng
    filtered_months = []
    filtered_trend1 = []
    filtered_trend2 = []

    for i, m in enumerate(months):
        if (i + 1) in months_to_show:
            filtered_months.append(m)
            filtered_trend1.append(trend1[i])
            filtered_trend2.append(trend2[i])

    if not filtered_months:
        print("⚠️ Không có tháng nào khớp với bộ lọc.")
        return

    # Load workbook
    book = load_workbook(file_path)
    if "Anomalies Summary" in book.sheetnames:
        sheet = book["Anomalies Summary"]
        start_row = sheet.max_row + 2
    else:
        sheet = book.create_sheet("Anomalies Summary")
        start_row = 1

    # Highlight styles
    diff_fill = PatternFill(start_color="FFFFCCCB", end_color="FFFFCCCB", fill_type="solid")
    red_font = Font(color="FF0000")

    # Header và dữ liệu
    header_row = [rule_name] + filtered_months + ["Remark"]
    row1 = [label1] + filtered_trend1 + [""]
    row2 = [label2] + filtered_trend2 + [""]

    # Tạo Remark cho sự khác biệt >5%
    remark_lines = []
    remark_text = ""
    for i in range(len(filtered_trend1)):
        t1 = filtered_trend1[i]
        t2 = filtered_trend2[i]
        print(f"Month {i + 1}: {t1} vs {t2}")

        if t1 != t2 and ("↑ >5%" in [t1, t2] or "↓ >5%" in [t1, t2] or "0" in [t1, t2]):
            month_num = i + 1
            remark_lines.append(f"Month {month_num}: Check BS vs PL trend")

    remark_text = "\n".join(remark_lines)
    row2[-1] = remark_text

    for row_offset, row_data in enumerate([header_row, row1, row2], start=1):
        for col_idx, val in enumerate(row_data, start=1):
            cell = sheet.cell(row=start_row + row_offset, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Tô màu nếu trend khác nhau
            if row_offset in [2, 3] and col_idx > 1:
                idx = col_idx - 2
                if idx < len(filtered_trend1) and filtered_trend1[idx] != filtered_trend2[idx]:
                    cell.fill = diff_fill
                    cell.font = red_font

    book.save(file_path)